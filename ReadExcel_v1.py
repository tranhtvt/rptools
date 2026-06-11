import pandas as pd
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

INPUT_FOLDER = "inputs"
OUTPUT_JSON = "output.json"
OUTPUT_EXCEL = "tong_hop.xlsx"

START_ROW = 6  # dòng 7

COLUMNS = [
    "soTT","tenUC","maTC","tacNhan","moTaUC","dkTruoc",
    "cacbuocKT","kqMongMuon","kqThucHien",
    "kqLan1","anhMC1","ghichuLan1","phanHoi1",
    "kqLan2","anhMC2","ghiChuLan2","phanHoi2"
]

def is_testcase(val):
    if pd.isna(val):
        return False
    return bool(re.match(r"\[.*?\]", str(val)))

# ================== READ ==================
all_data = []
sheet_blocks = []

for file in os.listdir(INPUT_FOLDER):
    if not file.endswith(".xlsx"):
        continue

    file_path = os.path.join(INPUT_FOLDER, file)
    excel = pd.ExcelFile(file_path)

    for sheet_index, sheet_name in enumerate(excel.sheet_names):
        if sheet_index == 0:
            continue

        df = excel.parse(sheet_name, header=None, dtype=str)

        # ===== LẤY HEADER NGAY DÒNG 7 =====
        header_row = df.iloc[START_ROW]

        soTT = "" if pd.isna(header_row[0]) else str(header_row[0]).strip()
        tenUC = "" if pd.isna(header_row[1]) else str(header_row[1]).strip()
        maTC = "" if pd.isna(header_row[2]) else str(header_row[2]).strip()
        tacNhan = "" if pd.isna(header_row[3]) else str(header_row[3]).strip()

        # xác định mã sheet để sort
        sheet_order = int(soTT) if soTT.isdigit() else 999

        block = {
            "order": sheet_order,
            "rows": []
        }

        # ===== ADD HEADER LỚN =====
        block["rows"].append({
            "type": "sheet_header",
            "value": [soTT, tenUC, maTC, tacNhan]
        })

        # ===== ĐỌC DATA BÊN DƯỚI =====
        df_data = df.iloc[START_ROW+1:].reset_index(drop=True)

        for _, row in df_data.iterrows():

            row_dict = {}
            for i, col in enumerate(COLUMNS):
                val = row[i] if i < len(row) else ""
                row_dict[col] = "" if pd.isna(val) else str(val).strip()

            # ===== TIÊU ĐỀ NHỎ =====
            if not is_testcase(row_dict["maTC"]):
                title = row_dict["tenUC"] or row_dict["soTT"]
                if title:
                    block["rows"].append({
                        "type": "title",
                        "value": title
                    })
                continue

            # ===== TEST CASE =====
            all_data.append(row_dict)
            block["rows"].append({
                "type": "data",
                "value": row_dict
            })

        sheet_blocks.append(block)

# ================== SORT THEO STT SHEET ==================
sheet_blocks = sorted(sheet_blocks, key=lambda x: x["order"])

# ================== EXPORT JSON ==================
with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

# ================== EXPORT EXCEL ==================
wb = Workbook()
ws = wb.active
ws.title = "TongHop"

# header
ws.append(COLUMNS)
row_idx = 2

for sheet in sheet_blocks:

    for item in sheet["rows"]:

        # ===== HEADER LỚN =====
        if item["type"] == "sheet_header":
            vals = item["value"]
            for col in range(4):
                ws.cell(row=row_idx, column=col+1, value=vals[col])
                ws.cell(row=row_idx, column=col+1).font = Font(bold=True)

            row_idx += 1
            continue

        # ===== TIÊU ĐỀ NHỎ =====
        if item["type"] == "title":
            ws.cell(row=row_idx, column=1, value=item["value"])
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1
            continue

        # ===== TEST CASE =====
        data = item["value"]
        ws.append([data.get(col, "") for col in COLUMNS])
        row_idx += 1

# wrap text
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

wb.save(OUTPUT_EXCEL)

print("✅ DONE: chuẩn 100% theo yêu cầu")