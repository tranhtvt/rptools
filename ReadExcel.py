import pandas as pd
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font

INPUT_FOLDER = "inputs"
OUTPUT_JSON = "output.json"
OUTPUT_EXCEL = "tong_hop.xlsx"
MAP_FILE = "maps.txt"

START_ROW = 7  # bắt đầu từ dòng 7 (index = 6)

COLUMNS = [
    "soTT","tenUC","maTC","tacNhan","moTaUC","dkTruoc",
    "cacbuocKT","kqMongMuon","kqThucHien",
    "kqLan1","anhMC1","ghichuLan1","phanHoi1",
    "kqLan2","anhMC2","ghiChuLan2","phanHoi2"
]

# ================== LOAD MAP ==================
def load_map():
    mapping = {}
    with open(MAP_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            match = re.match(r"(\d+)\.\s*(.+)", line)
            if match:
                code = match.group(1)
                name = match.group(2)
                mapping[code] = name
    return mapping

sheet_map = load_map()

# ================== CHECK TEST CASE ==================
def is_testcase(val):
    if pd.isna(val):
        return False
    return bool(re.match(r"\[.*?\]", str(val)))

# ================== READ DATA ==================
all_data = []
sheet_blocks = []

for file in os.listdir(INPUT_FOLDER):
    if not file.endswith(".xlsx"):
        continue

    file_path = os.path.join(INPUT_FOLDER, file)
    excel = pd.ExcelFile(file_path)

    for sheet_index, sheet_name in enumerate(excel.sheet_names):
        if sheet_index == 0:
            continue  # bỏ sheet đầu

        df = excel.parse(sheet_name, header=None, dtype=str)
        df = df.iloc[START_ROW:].reset_index(drop=True)

        # lấy mã sheet (ví dụ "1")
        sheet_code = re.findall(r"\d+", sheet_name)
        sheet_code = sheet_code[0] if sheet_code else sheet_name

        full_name = sheet_map.get(sheet_code, sheet_name)

        block_rows = []

        # dòng tiêu đề sheet
        block_rows.append({
            "type": "sheet",
            "value": f"{sheet_code}. {full_name}"
        })

        for _, row in df.iterrows():
            row_dict = {}
            for i, col in enumerate(COLUMNS):
                val = row[i] if i < len(row) else ""
                row_dict[col] = "" if pd.isna(val) else str(val)

            # ===== dòng tiêu đề (Thêm/Sửa/Xóa/...) =====
            if not is_testcase(row_dict["maTC"]):
                title = row_dict["tenUC"] or row_dict["soTT"]

                if title.strip():
                    block_rows.append({
                        "type": "title",
                        "value": title.strip()
                    })
                continue

            # ===== test case =====
            all_data.append(row_dict)
            block_rows.append({
                "type": "data",
                "value": row_dict
            })

        sheet_blocks.append({
            "sheet_code": int(sheet_code) if str(sheet_code).isdigit() else 999,
            "rows": block_rows
        })

# ================== SORT SHEET ==================
sheet_blocks = sorted(sheet_blocks, key=lambda x: x["sheet_code"])

# ================== EXPORT JSON ==================
with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

# ================== EXPORT EXCEL ==================
wb = Workbook()
ws = wb.active
ws.title = "TongHop"

# header
ws.append(COLUMNS)

row_idx = 2

for sheet in sheet_blocks:

    # ===== sheet title =====
    ws.cell(row=row_idx, column=1, value=sheet["rows"][0]["value"])
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    row_idx += 1

    for item in sheet["rows"][1:]:

        # ===== dòng tiêu đề (Thêm/Sửa/...) =====
        if item["type"] == "title":
            ws.cell(row=row_idx, column=1, value=item["value"])
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1
            continue

        # ===== dòng test case =====
        data = item["value"]
        ws.append([data.get(col, "") for col in COLUMNS])
        row_idx += 1

# auto wrap text
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrap_text=True)

wb.save(OUTPUT_EXCEL)

print("✅ DONE: JSON + Excel tổng hợp đã tạo")